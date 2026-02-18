import PyPDF2 
import docx 
import os 
from pathlib import Path 
from django .db .models import Q 
from .models import File 

def extract_text_from_pdf (file_path ):
    try :
        with open (file_path ,'rb')as file :
            pdf_reader =PyPDF2 .PdfReader (file )
            text =''
            for page_num in range (len (pdf_reader .pages )):
                page =pdf_reader .pages [page_num ]
                text +=page .extract_text ()or ''
            return text 
    except Exception as e :
        print (f"Error extracting text from PDF: {e }")
        return ""

def extract_text_from_txt (file_path ):
    try :
        with open (file_path ,'r',encoding ='utf-8')as file :
            return file .read ()
    except UnicodeDecodeError :
        try :
            with open (file_path ,'r',encoding ='latin-1')as file :
                return file .read ()
        except Exception as e :
            print (f"Error extracting text from TXT: {e }")
            return ""
    except Exception as e :
        print (f"Error extracting text from TXT: {e }")
        return ""

def extract_text_from_docx (file_path ):
    try :
        doc =docx .Document (file_path )
        text =''
        for paragraph in doc .paragraphs :
            text +=paragraph .text +'\n'
        return text 
    except Exception as e :
        print (f"Error extracting text from DOCX: {e }")
        return ""

def extract_text_from_xlsx (file_path ):
    try :
        import openpyxl 
        wb =openpyxl .load_workbook (file_path ,read_only =True )
        text =''
        for sheet_name in wb .sheetnames :
            sheet =wb [sheet_name ]
            text +=f"Sheet: {sheet_name }\n"
            for row in sheet .iter_rows (values_only =True ):
                text +=' | '.join (str (cell )if cell is not None else ''for cell in row )+'\n'
        return text 
    except Exception as e :
        print (f"Error extracting text from XLSX: {e }")
        return ""

def extract_text_from_pptx (file_path ):
    try :
        from pptx import Presentation 
        prs =Presentation (file_path )
        text =''
        for slide_number ,slide in enumerate (prs .slides ,1 ):
            text +=f"Slide {slide_number }:\n"
            for shape in slide .shapes :
                if hasattr (shape ,"text"):
                    text +=shape .text +'\n'
        return text 
    except Exception as e :
        print (f"Error extracting text from PPTX: {e }")
        return ""

def extract_text_from_file (file_path ):
    if not os .path .exists (file_path ):
        return ""

    ext =file_path .split ('.')[-1 ].lower ()

    if ext =='pdf':
        return extract_text_from_pdf (file_path )
    elif ext =='txt':
        return extract_text_from_txt (file_path )
    elif ext in ['docx','doc']:
        return extract_text_from_docx (file_path )
    elif ext in ['xlsx','xls']:
        return extract_text_from_xlsx (file_path )
    elif ext in ['pptx','ppt']:
        return extract_text_from_pptx (file_path )
    else :
        return ""

def generate_preview (file_path ,output_dir ):
    try :
        ext =file_path .split ('.')[-1 ].lower ()

        if ext in ['jpg','jpeg','png']:
            from PIL import Image 
            img =Image .open (file_path )
            img .thumbnail ((800 ,800 ))

            Path (output_dir ).mkdir (parents =True ,exist_ok =True )
            preview_path =os .path .join (output_dir ,'preview_'+os .path .basename (file_path ))
            img .save (preview_path )
            return preview_path 

    except Exception as e :
        print (f"Error generating preview: {e }")
        return None 

    return None 

def search_files (query ,user =None ,file_types =None ,categories =None ,tags =None ,date_from =None ,date_to =None ):
    files =File .objects .filter (is_folder =False )

    if user :
        files =files .filter (
        Q (uploaded_by =user )|
        Q (visibility ='public')|
        Q (shared_with =user )
        ).distinct ()

    if query :
        files =files .filter (
        Q (title__icontains =query )|
        Q (description__icontains =query )|
        Q (extracted_text__icontains =query )
        )

    if file_types :
        files =files .filter (file_type__in =file_types )

    if categories :
        files =files .filter (category__in =categories )

    if tags :
        for tag in tags :
            files =files .filter (tags =tag )

    if date_from :
        files =files .filter (uploaded_at__date__gte =date_from )

    if date_to :
        files =files .filter (uploaded_at__date__lte =date_to )

    return files .distinct ()

def get_user_storage_usage (user ):
    from .models import UserStorageQuota 

    quota ,created =UserStorageQuota .objects .get_or_create (
    user =user ,
    defaults ={'total_quota_bytes':5368709120 }
    )

    if created :
        quota .update_usage ()

    return quota 