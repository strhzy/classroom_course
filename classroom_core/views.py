from django.shortcuts import render ,redirect ,get_object_or_404 
from django.urls import reverse
from django.contrib.auth.decorators import login_required 
from django.contrib import messages 
from django.core.exceptions import PermissionDenied 
from django.db.models import Q 
from django.core.paginator import Paginator 
from django.contrib.auth.models import User 
from django.http import JsonResponse
from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
import io
import csv
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from file_manager.models import File
from file_manager.views import create_user_uploaded_file
from . models import *
from . forms import *
from django.utils import timezone 
from web_messages import flash_form_errors


WEEKDAY_MAP = {
    "пн": 0, "mon": 0, "monday": 0,
    "вт": 1, "tue": 1, "tuesday": 1,
    "ср": 2, "wed": 2, "wednesday": 2,
    "чт": 3, "thu": 3, "thursday": 3,
    "пт": 4, "fri": 4, "friday": 4,
    "сб": 5, "sat": 5, "saturday": 5,
    "вс": 6, "sun": 6, "sunday": 6,
}
WEEKDAY_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


def _parse_class_days(class_days_raw):
    selected = set()
    for token in (class_days_raw or "").replace(" ", "").lower().split(","):
        if token in WEEKDAY_MAP:
            selected.add(WEEKDAY_MAP[token])
    return selected


def _serialize_assignment_quiz(assignment):
    result = []
    for question in assignment.quiz_questions.prefetch_related("options").all():
        result.append(
            {
                "text": question.question_text,
                "options": [
                    {"text": option.option_text, "is_correct": option.is_correct}
                    for option in question.options.all()
                ],
            }
        )
    return result


def _save_assignment_quiz(assignment, quiz_payload):
    assignment.quiz_questions.all().delete()
    for question_idx, question_data in enumerate(quiz_payload):
        question_text = (question_data.get("text") or "").strip()
        if not question_text:
            continue
        question_obj = AssignmentQuizQuestion.objects.create(
            assignment=assignment,
            question_text=question_text,
            order=question_idx,
        )
        options = question_data.get("options") or []
        for option_idx, option_data in enumerate(options):
            option_text = (option_data.get("text") or "").strip()
            if not option_text:
                continue
            AssignmentQuizOption.objects.create(
                question=question_obj,
                option_text=option_text,
                is_correct=bool(option_data.get("is_correct")),
                order=option_idx,
            )


def _ensure_year_schedule(course):
    if not course.start_date:
        return
    selected_weekdays = _parse_class_days(course.class_days)
    if not selected_weekdays:
        return
    start = course.start_date
    finish = course.end_date or (start + timedelta(days=365))
    existing_lessons = CourseLesson.objects.filter(course=course)
    existing_lessons.exclude(lesson_date__range=(start, finish)).delete()
    existing_lessons.exclude(lesson_date__week_day__in=[((day + 1) % 7) + 1 for day in selected_weekdays]).delete()
    current = start
    while current <= finish:
        if current.weekday() in selected_weekdays:
            CourseLesson.objects.get_or_create(
                course=course,
                lesson_date=current,
                lesson_number=1,
            )
        current += timedelta(days=1)

@login_required 
def course_list(request ):
    """Список курсов"""
    user_profile = request.user.profile 

    if request.user.is_superuser :
        courses = Course.objects.all()
    else :
        courses = Course.objects.filter(
        Q(instructor = request.user )|
        Q(teaching_assistants = request.user )|
        Q(students = request.user )
        ).distinct()


    status =request.GET.get('status')
    if status :
        courses =courses.filter(status =status )


    query =request.GET.get('query')
    if query :
        courses =courses.filter(
        Q(title__icontains =query )|
        Q(description__icontains =query )|
        Q(code__icontains =query )
        )


    sort_by =request.GET.get('sort','-created_at')
    courses =courses.order_by(sort_by )


    paginator =Paginator(courses ,12 )
    page_number =request.GET.get('page')
    page_obj =paginator.get_page(page_number )

    context ={
    'page_obj':page_obj ,
    'status':status ,
    'query':query ,
    'sort_by':sort_by ,
    }

    return render(request ,'classroom_core/course_list.html',context )

@login_required
def course_detail(request ,course_id ):
    """Детали курса"""
    course =get_object_or_404(Course ,id =course_id )


    if not course.can_access(request.user ):
        raise PermissionDenied


    is_student =course.students.filter(id =request.user.id ).exists()
    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all()
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())
    
                                                                                   
    can_manage_course = is_teacher or is_assistant or is_admin

    announcements = course.announcements.all().order_by('-is_pinned', '-created_at')
    sections = course.sections.prefetch_related('materials').all().order_by('order')
    if not can_manage_course:
        sections = sections.filter(is_visible=True)
    assignments =course.assignments.filter(status ='published').order_by('-due_date')
    materials_qs = CourseMaterial.objects.filter(
        section__course=course,
        status="published"
    ).select_related("section").order_by("-created_at")

    if not can_manage_course:
        materials_qs = materials_qs.filter(is_visible=True, section__is_visible=True)

    stream_items = []
    for announcement in announcements:
        stream_items.append({
            "type": "announcement",
            "title": announcement.title,
            "description": announcement.content,
            "section_title": None,
            "date": announcement.created_at,
            "meta": announcement.author.get_full_name() or announcement.author.username,
            "announcement_id": announcement.id,
            "is_pinned": announcement.is_pinned,
        })

    for material in materials_qs:
        stream_items.append({
            "type": "material",
            "title": material.title,
            "description": material.description or material.content,
            "section_title": material.section.title,
            "date": material.created_at,
            "meta": material.get_material_type_display(),
            "url": None,
        })

    for assignment in assignments:
        stream_items.append({
            "type": "assignment",
            "title": assignment.title,
            "description": assignment.description,
            "section_title": None,
            "date": assignment.created_at,
            "meta": f"{assignment.max_points} баллов",
            "assignment_id": assignment.id,
        })

    stream_items.sort(key=lambda item: item["date"] or timezone.now(), reverse=True)
    students = course.get_all_enrolled_students()
    current_assistants = course.teaching_assistants.select_related("profile").all()
    available_teachers = User.objects.filter(profile__role='teacher').exclude(
        id__in=list(current_assistants.values_list("id", flat=True)) + [course.instructor_id]
    ).order_by("last_name", "first_name", "username")

                                                    
    user_has_enrollment_request = False
    user_enrollment_request_id = None
    if not is_student and not can_manage_course:
        enrollment_request = CourseEnrollmentRequest.objects.filter(
            course=course,
            student=request.user
        ).first()
        if enrollment_request:
            user_has_enrollment_request = True
            user_enrollment_request_id = enrollment_request.id

                                               
    student_submissions = []
    if is_student:
        student_submissions = AssignmentSubmission.objects.filter(
            student=request.user,
            assignment__course=course
        ).select_related('assignment')[:5]
        student_submissions_count = AssignmentSubmission.objects.filter(
            student=request.user,
            assignment__course=course
        ).count()

    context ={
    'course':course ,
    'announcements':announcements ,
    'sections':sections ,
    'assignments':assignments ,
    'stream_items': stream_items,
    'students':students ,
    'current_assistants': current_assistants,
    'available_teachers': available_teachers,
    'is_student':is_student ,
    'is_teacher': is_teacher,
    'is_assistant': is_assistant,
    'is_admin': is_admin,
    'can_manage_course': can_manage_course,
    'user_has_enrollment_request': user_has_enrollment_request,
    'user_enrollment_request_id': user_enrollment_request_id,
    'student_submissions': student_submissions,
    'student_submissions_count': student_submissions_count if is_student else 0,
    }

    return render(request ,'classroom_core/course_detail.html',context )


@login_required
@require_http_methods(["POST"])
def course_manage_teaching_assistants(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if not course.can_edit(request.user):
        raise PermissionDenied

    action = request.POST.get("action", "add")
    teacher_id = request.POST.get("teacher_id")
    if not teacher_id:
        messages.error(request, "Не выбран преподаватель")
        return redirect("classroom_core:course_detail", course_id=course.id)

    teacher = User.objects.filter(id=teacher_id, profile__role="teacher").first()
    if not teacher:
        messages.error(request, "Преподаватель не найден")
        return redirect("classroom_core:course_detail", course_id=course.id)

    if action == "remove":
        course.teaching_assistants.remove(teacher)
        messages.success(request, "Преподаватель удален из ассистентов")
    else:
        if teacher == course.instructor:
            messages.info(request, "Этот пользователь уже основной преподаватель курса")
        else:
            course.teaching_assistants.add(teacher)
            messages.success(request, "Преподаватель добавлен в ассистенты курса")

    return redirect("classroom_core:course_detail", course_id=course.id)


@login_required
def course_submissions(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if not course.can_edit(request.user):
        raise PermissionDenied

    submissions = AssignmentSubmission.objects.filter(
        assignment__course=course
    ).select_related("assignment", "student", "graded_by").order_by("-submitted_at", "-updated_at")

    return render(request, "classroom_core/course_submissions.html", {
        "course": course,
        "submissions": submissions,
    })

@login_required
def course_create(request):
    """Создание нового курса"""
    user_profile = request.user.profile
    
                                                             
    if not (user_profile.is_teacher() or user_profile.is_staff() or request.user.is_superuser):
        raise PermissionDenied
    
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            course = form.save(commit=False)
            course.instructor = request.user
            if course.start_date and not course.end_date:
                course.end_date = course.start_date + timedelta(days=365)
            course.save()
            form.save_m2m()
            _ensure_year_schedule(course)
            
            messages.success(request, 'Курс успешно создан. Чат курса автоматически создан для всех участников.')
            return redirect('classroom_core:course_detail', course_id=course.id)
        else:
            flash_form_errors(request, form)
    else:
        form = CourseForm(user=request.user)
    
    return render(request, 'classroom_core/course_form.html', {
        'form': form,
        'title': 'Создать курс'
    })

@login_required 
def course_edit(request ,course_id ):
    """Редактирование курса"""
    course =get_object_or_404(Course ,id =course_id )


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =CourseForm(request.POST ,request.FILES ,instance =course )
        if form.is_valid():
            updated_course = form.save()
            _ensure_year_schedule(updated_course)
            messages.success(request ,'Курс успешно обновлен')
            return redirect('classroom_core:course_detail',course_id =course.id )
        else:
            flash_form_errors(request, form)
    else :
        form =CourseForm(instance =course )

    return render(request ,'classroom_core/course_form.html',{
    'form':form ,
    'title':'Редактировать курс',
    'course':course 
    })

@login_required 
def course_delete(request ,course_id ):
    """Удаление курса"""
    course =get_object_or_404(Course ,id =course_id )


    if not course.can_delete(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        course.delete()
        messages.success(request ,'Курс успешно удален')
        return redirect('classroom_core:course_list')

    return render(request ,'classroom_core/course_confirm_delete.html',{
    'course':course 
    })



@login_required 
def section_create(request ,course_id ):
    """Создание раздела курса"""
    course =get_object_or_404(Course ,id =course_id )


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =CourseSectionForm(request.POST )
        if form.is_valid():
            section =form.save(commit =False )
            section.course =course 
            section.save()
            messages.success(request ,'Раздел успешно создан')
            return redirect('classroom_core:course_detail',course_id =course.id )
        else:
            flash_form_errors(request, form)
    else :
        form =CourseSectionForm()

    return render(request ,'classroom_core/section_form.html',{
    'form':form ,
    'course':course ,
    'title':'Создать раздел'
    })

@login_required 
def section_edit(request ,section_id ):
    """Редактирование раздела курса"""
    section =get_object_or_404(CourseSection ,id =section_id )
    course =section.course 


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =CourseSectionForm(request.POST ,instance =section )
        if form.is_valid():
            form.save()
            messages.success(request ,'Раздел успешно обновлен')
            return redirect('classroom_core:course_detail',course_id =course.id )
        else:
            flash_form_errors(request, form)
    else :
        form =CourseSectionForm(instance =section )

    return render(request ,'classroom_core/section_form.html',{
    'form':form ,
    'course':course ,
    'section':section ,
    'title':'Редактировать раздел'
    })

@login_required 
def section_delete(request ,section_id ):
    """Удаление раздела курса"""
    section =get_object_or_404(CourseSection ,id =section_id )
    course =section.course 


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        section.delete()
        messages.success(request ,'Раздел успешно удален')
        return redirect('classroom_core:course_detail',course_id =course.id )

    return render(request ,'classroom_core/section_confirm_delete.html',{
    'section':section ,
    'course':course 
    })



@login_required 
def material_create(request ,section_id ):
    """Создание учебного материала"""
    section =get_object_or_404(CourseSection ,id =section_id )
    course =section.course 


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =CourseMaterialForm(request.POST ,request.FILES )
        if form.is_valid():
            material =form.save(commit =False )
            material.section =section 
            material.save()
            messages.success(request ,'Материал успешно создан')
            return redirect('classroom_core:course_detail',course_id =course.id )
        else:
            flash_form_errors(request, form)
    else :
        form =CourseMaterialForm()

    return render(request ,'classroom_core/material_form.html',{
    'form':form ,
    'section':section ,
    'course':course ,
    'title':'Создать материал'
    })

@login_required 
def material_edit(request ,material_id ):
    """Редактирование учебного материала"""
    material =get_object_or_404(CourseMaterial ,id =material_id )
    section =material.section 
    course =section.course 


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =CourseMaterialForm(request.POST ,request.FILES ,instance =material )
        if form.is_valid():
            form.save()
            messages.success(request ,'Материал успешно обновлен')
            return redirect('classroom_core:course_detail',course_id =course.id )
        else:
            flash_form_errors(request, form)
    else :
        form =CourseMaterialForm(instance =material )

    return render(request ,'classroom_core/material_form.html',{
    'form':form ,
    'section':section ,
    'course':course ,
    'material':material ,
    'title':'Редактировать материал'
    })

@login_required 
def material_delete(request ,material_id ):
    """Удаление учебного материала"""
    material =get_object_or_404(CourseMaterial ,id =material_id )
    section =material.section 
    course =section.course 


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        material.delete()
        messages.success(request ,'Материал успешно удален')
        return redirect('classroom_core:course_detail',course_id =course.id )

    return render(request ,'classroom_core/material_confirm_delete.html',{
    'material':material ,
    'section':section ,
    'course':course 
    })



@login_required 
def assignment_list(request ):
    """Список заданий"""
    user_profile =request.user.profile 
    course_queryset = Course.objects.filter(
        Q(instructor=request.user) |
        Q(teaching_assistants=request.user) |
        Q(students=request.user)
    ).distinct().order_by("title")

    if user_profile.is_teacher() or user_profile.is_staff() or request.user.is_superuser:

        assignments = Assignment.objects.all()
        if not (user_profile.is_staff() or request.user.is_superuser):
            assignments = assignments.filter(
                Q(course__instructor=request.user) |
                Q(course__teaching_assistants=request.user) |
                Q(course__students=request.user)
            ).distinct()
    else :

        assignments =Assignment.objects.filter(
        course__students =request.user ,
        status ='published'
        )


    course_id = request.GET.get('course_id')
    selected_course_id = None
    if course_id:
        try:
            selected_course_id = int(course_id)
        except (TypeError, ValueError):
            selected_course_id = None
        if selected_course_id and course_queryset.filter(id=selected_course_id).exists():
            assignments = assignments.filter(course_id=selected_course_id)


    status =request.GET.get('status')
    if status:
        assignments=assignments.filter(status =status )


    sort_by =request.GET.get('sort','-due_date')
    assignments =assignments.order_by(sort_by )


    paginator =Paginator(assignments ,12 )
    page_number =request.GET.get('page')
    page_obj =paginator.get_page(page_number )

    context ={
    'page_obj':page_obj ,
    'course_id':selected_course_id ,
    'status':status ,
    'sort_by':sort_by ,
    'filter_courses': course_queryset,
    }

    return render(request ,'classroom_core/assignment_list.html',context )

@login_required
def assignment_detail(request ,assignment_id ):
    """Детали задания"""
    assignment =get_object_or_404(Assignment ,id =assignment_id )
    course =assignment.course


    if not course.can_access(request.user ):
        raise PermissionDenied


    is_student =course.students.filter(id =request.user.id ).exists()
    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all()
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())
    
                                                  
    can_manage_course = is_teacher or is_assistant or is_admin

    submission =None
    if is_student :
        submission =AssignmentSubmission.objects.filter(
        assignment =assignment ,
        student =request.user
        ).first()

    latest_quiz_attempt = None
    if assignment.assignment_type == "quiz" and is_student:
        latest_quiz_attempt = AssignmentQuizAttempt.objects.filter(
            assignment=assignment,
            student=request.user,
        ).order_by("-submitted_at").first()
    
                                                   
    all_submissions = []
    if can_manage_course:
        all_submissions = list(AssignmentSubmission.objects.filter(
            assignment=assignment
        ).select_related('student', 'graded_by').order_by('-submitted_at'))
    
                           
    submissions_stats = {}
    quiz_results_by_student = {}
    if can_manage_course:
        total = len(all_submissions)
        graded = sum(1 for sub in all_submissions if sub.status == "graded")
        submitted = sum(1 for sub in all_submissions if sub.status == "submitted")
        returned = sum(1 for sub in all_submissions if sub.status == "returned")
        late = (
            sum(1 for sub in all_submissions if sub.submitted_at and sub.submitted_at > assignment.due_date)
            if assignment.due_date
            else 0
        )
        
        submissions_stats = {
            'total': total,
            'graded': graded,
            'submitted': submitted,
            'returned': returned,
            'late': late,
        }
        latest_attempts = {}
        for attempt in AssignmentQuizAttempt.objects.filter(assignment=assignment).select_related("student").order_by("student_id", "-submitted_at"):
            if attempt.student_id not in latest_attempts:
                latest_attempts[attempt.student_id] = attempt
        quiz_results_by_student = {
            student_id: f"{attempt.correct_answers}/{attempt.total_questions}"
            for student_id, attempt in latest_attempts.items()
        }
        for sub in all_submissions:
            sub.quiz_correctness = quiz_results_by_student.get(sub.student_id, "0/0")

    context ={
    'assignment':assignment ,
    'course':course ,
    'is_student':is_student ,
    'is_teacher': is_teacher,
    'is_assistant': is_assistant,
    'is_admin': is_admin,
    'can_manage_course': can_manage_course,
    'submission':submission ,
    'all_submissions': all_submissions,
    'submissions_stats': submissions_stats,
    'student_files': AssignmentFile.objects.filter(assignment=assignment, student=request.user).order_by('-uploaded_at') if is_student else [],
    'quiz_questions': assignment.quiz_questions.prefetch_related("options").all() if assignment.assignment_type == "quiz" else [],
    "latest_quiz_attempt": latest_quiz_attempt,
    "quiz_results_by_student": quiz_results_by_student,
    }

    return render(request ,'classroom_core/assignment_detail.html',context )

@login_required 
def assignment_create(request ,course_id ):
    """Создание задания"""
    course = get_object_or_404(Course, id=course_id )


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form = AssignmentForm(request.POST, request.FILES )
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.course = course
            assignment.save()
            if assignment.assignment_type == "quiz":
                quiz_payload_raw = request.POST.get("quiz_payload", "[]")
                try:
                    quiz_payload = json.loads(quiz_payload_raw)
                except (TypeError, ValueError):
                    quiz_payload = []
                _save_assignment_quiz(assignment, quiz_payload)
            else:
                assignment.quiz_questions.all().delete()
            messages.success(request ,'Задание успешно создано')
            return redirect('classroom_core:course_detail',course_id=course.id)
        else:
            flash_form_errors(request, form)
    else:
        form = AssignmentForm()

    return render(request ,'classroom_core/assignment_form.html',{
    'form':form ,
    'course':course ,
    'title':'Создать задание',
    "quiz_initial_json": json.dumps([], ensure_ascii=False),
    })

@login_required 
def assignment_edit(request ,assignment_id ):
    """Редактирование задания"""
    assignment =get_object_or_404(Assignment ,id =assignment_id )
    course =assignment.course 


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =AssignmentForm(request.POST ,request.FILES ,instance =assignment )
        if form.is_valid():
            updated_assignment = form.save()
            if updated_assignment.assignment_type == "quiz":
                quiz_payload_raw = request.POST.get("quiz_payload", "[]")
                try:
                    quiz_payload = json.loads(quiz_payload_raw)
                except (TypeError, ValueError):
                    quiz_payload = []
                _save_assignment_quiz(updated_assignment, quiz_payload)
            else:
                updated_assignment.quiz_questions.all().delete()
            messages.success(request ,'Задание успешно обновлено')
            return redirect('classroom_core:assignment_detail',assignment_id =assignment.id )
        else:
            flash_form_errors(request, form)
    else :
        form =AssignmentForm(instance =assignment )

    return render(request ,'classroom_core/assignment_form.html',{
    'form':form ,
    'course':course ,
    'assignment':assignment ,
    'title':'Редактировать задание',
    "quiz_initial_json": json.dumps(_serialize_assignment_quiz(assignment), ensure_ascii=False),
    })

@login_required 
def assignment_delete(request ,assignment_id ):
    """Удаление задания"""
    assignment =get_object_or_404(Assignment ,id =assignment_id )
    course =assignment.course 


    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        assignment.delete()
        messages.success(request ,'Задание успешно удалено')
        return redirect('classroom_core:course_detail',course_id =course.id )

    return render(request ,'classroom_core/assignment_confirm_delete.html',{
    'assignment':assignment ,
    'course':course 
    })

@login_required 
def assignment_submit(request ,assignment_id ):
    """Отправка решения задания (текст + файл с устройства или из хранилища)."""
    assignment = get_object_or_404(Assignment, id=assignment_id)
    course = assignment.course

    if assignment.assignment_type == "quiz":
        messages.info(request, "Для этого задания ответы отправляются в форме на странице задания.")
        return redirect("classroom_core:assignment_detail", assignment_id=assignment.id)

    if not course.can_access(request.user ):
        raise PermissionDenied 

    if not course.students.filter(id =request.user.id ).exists():
        raise PermissionDenied 

    if not assignment.can_submit():
        messages.error(request ,'Отправка решений для этого задания невозможна')
        return redirect('classroom_core:assignment_detail',assignment_id =assignment.id )

    existing_submission = AssignmentSubmission.objects.filter(
        assignment=assignment,
        student=request.user,
    ).first()

    if existing_submission and existing_submission.status !='returned':
        messages.error(request ,'Вы уже отправили решение этого задания')
        return redirect('classroom_core:assignment_detail',assignment_id =assignment.id )

    if request.method == 'POST':
        form = AssignmentSubmitCombinedForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            desc = form.cleaned_data.get('attachment_description') or ''
            text = (form.cleaned_data.get('text_response') or '').strip()
            upload = form.cleaned_data.get('upload_from_pc')
            storage = form.cleaned_data.get('storage_file')

            file_record = None
            upload_scan = None
            if upload:
                file_record, err, upload_scan = create_user_uploaded_file(request.user, upload)
                if err:
                    messages.error(request, err)
                    return render(request, 'classroom_core/assignment_submit.html', {
                        'form': form,
                        'assignment': assignment,
                        'course': course,
                    })
            elif storage:
                file_record = storage

            if existing_submission:
                submission = existing_submission
            else:
                submission = AssignmentSubmission(assignment=assignment, student=request.user)

            submission.text_response = text
            if existing_submission and existing_submission.status == "returned":
                submission.status = "submitted"
            submission.save()

            if file_record:
                AssignmentFile.objects.create(
                    assignment=assignment,
                    student=request.user,
                    file=file_record,
                    description=desc,
                )

            messages.success(request ,'Решение успешно отправлено')
            if upload and upload_scan:
                from file_manager.clamav import flash_scan_followup
                flash_scan_followup(request, upload_scan)
            return redirect('classroom_core:assignment_detail',assignment_id =assignment.id )
        else:
            flash_form_errors(request, form)
    else :
        initial = {}
        if existing_submission:
            initial['text_response'] = existing_submission.text_response or ''
        form = AssignmentSubmitCombinedForm(initial=initial, user=request.user)

    return render(request ,'classroom_core/assignment_submit.html',{
    'form':form ,
    'assignment':assignment ,
    'course':course 
    })

@login_required 
def assignment_grade(request ,submission_id ):
    """Оценка решения задания"""
    submission =get_object_or_404(AssignmentSubmission ,id =submission_id )
    assignment =submission.assignment 
    course =assignment.course 


    if not assignment.can_grade(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =AssignmentGradeForm(request.POST ,instance =submission )
        if form.is_valid():
            submission =form.save(commit =False )
            submission.graded_by =request.user 
            submission.graded_at =timezone.now()
            submission.save()

            messages.success(request ,'Решение успешно оценено')
            return redirect('classroom_core:assignment_detail',assignment_id =assignment.id )
        else:
            flash_form_errors(request, form)
    else :
        form =AssignmentGradeForm(instance =submission )

    submission_assignment_files = AssignmentFile.objects.filter(
        assignment=assignment,
        student=submission.student,
    ).select_related("file").order_by("-uploaded_at")

    return render(request ,'classroom_core/assignment_grade.html',{
    'form':form ,
    'submission':submission ,
    'assignment':assignment ,
    'course':course ,
    'submission_assignment_files': submission_assignment_files,
    })



@login_required 
def announcement_list(request ):
    """Список объявлений"""
    user_profile =request.user.profile 
    course_queryset = Course.objects.filter(
        Q(instructor=request.user) |
        Q(teaching_assistants=request.user) |
        Q(students=request.user)
    ).distinct().order_by("title")

    if user_profile.is_teacher()or user_profile.is_staff()or request.user.is_superuser :

        announcements = Announcement.objects.all()
        if not (user_profile.is_staff() or request.user.is_superuser):
            announcements = announcements.filter(
                Q(course__instructor=request.user) |
                Q(course__teaching_assistants=request.user) |
                Q(course__students=request.user)
            ).distinct()
    else :

        announcements =Announcement.objects.filter(
        course__students =request.user 
        )


    course_id = request.GET.get('course_id')
    selected_course_id = None
    if course_id:
        try:
            selected_course_id = int(course_id)
        except (TypeError, ValueError):
            selected_course_id = None
        if selected_course_id and course_queryset.filter(id=selected_course_id).exists():
            announcements = announcements.filter(course_id=selected_course_id)


    sort_by =request.GET.get('sort','-is_pinned')
    announcements =announcements.order_by(sort_by )

    paginator =Paginator(announcements ,12 )
    page_number =request.GET.get('page')
    page_obj =paginator.get_page(page_number )

    context ={
    'page_obj':page_obj ,
    'course_id':selected_course_id ,
    'sort_by':sort_by ,
    'filter_courses': course_queryset,
    }

    return render(request ,'classroom_core/announcement_list.html',context )

@login_required 
def announcement_detail(request ,announcement_id ):
    announcement =get_object_or_404(Announcement ,id =announcement_id )
    course =announcement.course 

    if not course.can_access(request.user ):
        raise PermissionDenied 

    return render(request ,'classroom_core/announcement_detail.html',{
    'announcement':announcement ,
    'course':course 
    })

@login_required 
def announcement_create(request ,course_id ):
    course =get_object_or_404(Course ,id =course_id )

    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =AnnouncementForm(request.POST )
        if form.is_valid():
            announcement =form.save(commit =False )
            announcement.course =course 
            announcement.author =request.user 
            announcement.save()
            messages.success(request ,'Объявление успешно создано')
            return redirect('classroom_core:course_detail',course_id =course.id )
        else:
            flash_form_errors(request, form)
    else :
        form =AnnouncementForm()

    return render(request ,'classroom_core/announcement_form.html',{
    'form':form ,
    'course':course ,
    'title':'Создать объявление'
    })

@login_required 
def announcement_edit(request ,announcement_id ):
    announcement =get_object_or_404(Announcement ,id =announcement_id )
    course =announcement.course 

    if not announcement.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =AnnouncementForm(request.POST ,instance =announcement )
        if form.is_valid():
            form.save()
            messages.success(request ,'Объявление успешно обновлено')
            return redirect('classroom_core:announcement_detail',announcement_id =announcement.id )
        else:
            flash_form_errors(request, form)
    else :
        form =AnnouncementForm(instance =announcement )

    return render(request ,'classroom_core/announcement_form.html',{
    'form':form ,
    'course':course ,
    'announcement':announcement ,
    'title':'Редактировать объявление'
    })

@login_required 
def announcement_delete(request ,announcement_id ):
    announcement =get_object_or_404(Announcement ,id =announcement_id )
    course =announcement.course 

    if not announcement.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        announcement.delete()
        messages.success(request ,'Объявление успешно удалено')
        return redirect('classroom_core:course_detail',course_id =course.id )

    return render(request ,'classroom_core/announcement_confirm_delete.html',{
    'announcement':announcement ,
    'course':course 
    })

@login_required 
def student_list(request ,course_id ):
    course =get_object_or_404(Course ,id =course_id )

    if not course.can_edit(request.user ):
        raise PermissionDenied 

    students =course.students.all()

    query =request.GET.get('query')
    if query :
        students =students.filter(
        Q(username__icontains =query )|
        Q(first_name__icontains =query )|
        Q(last_name__icontains =query )|
        Q(email__icontains =query )
        )

    paginator =Paginator(students ,20 )
    page_number =request.GET.get('page')
    page_obj =paginator.get_page(page_number )

    context ={
    'course':course ,
    'page_obj':page_obj ,
    'query':query ,
    }

    return render(request ,'classroom_core/student_list.html',context )

@login_required 
def student_enroll(request ,course_id ):
    course =get_object_or_404(Course ,id =course_id )

    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        form =StudentEnrollmentForm(request.POST )
        if form.is_valid():
            students =form.cleaned_data ['students']
            for student in students :
                success ,message =course.add_student(student )
                if not success :
                    messages.warning(request ,f'{student.username }: {message }')

            if students :
                messages.success(request ,'Студенты успешно зачислены на курс')
            return redirect('classroom_core:student_list',course_id =course.id )
        else:
            flash_form_errors(request, form)
    else :
        form =StudentEnrollmentForm()

    return render(request ,'classroom_core/student_enroll.html',{
    'form':form ,
    'course':course 
    })

@login_required 
def student_remove(request ,course_id ,student_id ):
    course =get_object_or_404(Course ,id =course_id )
    student =get_object_or_404(User ,id =student_id )

    if not course.can_edit(request.user ):
        raise PermissionDenied 

    if request.method =='POST':
        success ,message =course.remove_student(student )
        if success :
            messages.success(request ,'Студент успешно удален с курса')
        else :
            messages.error(request ,message )
        return redirect('classroom_core:student_list',course_id =course.id )

    return render(request ,'classroom_core/student_confirm_remove.html',{
    'course':course ,
    'student':student 
    })


@login_required 
def profile_view(request ,user_id =None ):
    if user_id :
        user =get_object_or_404(User ,id =user_id )
    else :
        user =request.user 

    return render(request ,'classroom_core/profile_view.html',{
    'profile_user':user 
    })

@login_required 
def profile_edit(request ):
    if request.method =='POST':
        form =UserProfileForm(request.POST ,request.FILES ,instance =request.user.profile )
        if form.is_valid():
            form.save()
            messages.success(request ,'Профиль успешно обновлен')
            return redirect('classroom_core:profile_view')
        else:
            flash_form_errors(request, form)
    else :
        form =UserProfileForm(instance =request.user.profile )

    return render(request ,'classroom_core/profile_form.html',{
    'form':form ,
    'title':'Редактировать профиль'
    })

@login_required
def student_list(request, course_id):
    """Список студентов курса"""
    course = get_object_or_404(Course, id=course_id)
    
                   
    if not course.can_edit(request.user):
        raise PermissionDenied
    
                                                              
    all_students = course.get_all_enrolled_students()
    
           
    query = request.GET.get('query')
    if query:
        all_students = all_students.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
               
    paginator = Paginator(list(all_students), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'course': course,
        'page_obj': page_obj,
        'query': query,
        'individual_students': course.students.all(),
        'group_students': course.student_groups.all(),
    }
    
    return render(request, 'classroom_core/student_list.html', context)

@login_required
def student_enroll(request, course_id):
    """Зачисление студентов и групп на курс"""
    course = get_object_or_404(Course, id=course_id)

                   
    if not course.can_edit(request.user):
        raise PermissionDenied

    if request.method == 'POST':
                                                              
        student_ids = request.POST.getlist('students')
        group_ids = request.POST.getlist('groups')

        students_added = 0
        groups_added = 0

                                            
        for student_id in student_ids:
            try:
                student = User.objects.get(id=student_id)
                success, message = course.add_student(student)
                if success:
                    students_added += 1
                else:
                    messages.warning(request, f'{student.username}: {message}')
            except User.DoesNotExist:
                messages.error(request, 'Студент не найден')

                                   
        for group_id in group_ids:
            try:
                group = StudentGroup.objects.get(id=group_id)
                success, message = course.add_student_group(group)
                if success:
                    groups_added += 1
                else:
                    messages.warning(request, f'{group.name}: {message}')
            except StudentGroup.DoesNotExist:
                messages.error(request, 'Группа не найдена')

        if students_added > 0 or groups_added > 0:
            messages.success(request, f'Успешно зачислено: {students_added} студентов и {groups_added} групп')
        else:
            messages.warning(request, 'Никто не был зачислен на курс')
        
        return redirect('classroom_core:student_list', course_id=course.id)

                                   
                                                 
    available_students = User.objects.filter(
        profile__role='student'
    ).exclude(
        id__in=course.students.all()
    )
    
                                              
    available_groups = StudentGroup.objects.exclude(
        id__in=course.student_groups.all()
    )

    context = {
        'course': course,
        'available_students': available_students,
        'available_groups': available_groups,
    }

    return render(request, 'classroom_core/student_enroll.html', context)

@login_required
def student_remove(request, course_id, student_id):
    """Удаление студента с курса"""
    course = get_object_or_404(Course, id=course_id)
    student = get_object_or_404(User, id=student_id)
    
                   
    if not course.can_edit(request.user):
        raise PermissionDenied
    
    if request.method == 'POST':
        success, message = course.remove_student(student)
        if success:
            messages.success(request, 'Студент успешно удален с курса')
        else:
            messages.error(request, message)
        return redirect('classroom_core:student_list', course_id=course.id)
    
    return render(request, 'classroom_core/student_confirm_remove.html', {
        'course': course,
        'student': student
    })

@login_required
def group_list(request):
    """Список групп студентов"""
    if not(request.user.profile.is_teacher() or request.user.profile.is_staff() or request.user.is_superuser):
        raise PermissionDenied
    
    groups = StudentGroup.objects.all()
    
           
    query = request.GET.get('query')
    if query:
        groups = groups.filter(
            Q(name__icontains=query) |
            Q(description__icontains=query)
        )
    
               
    paginator = Paginator(groups, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
    }
    
    return render(request, 'classroom_core/group_list.html', context)

@login_required
def group_create(request):
    """Создание группы студентов"""
    if not(request.user.profile.is_teacher() or request.user.profile.is_staff() or request.user.is_superuser):
        raise PermissionDenied
    
    if request.method == 'POST':
        form = StudentGroupForm(request.POST)
        if form.is_valid():
            group = form.save(commit=False)
            group.created_by = request.user
            group.save()
            messages.success(request, 'Группа успешно создана')
            return redirect('classroom_core:group_list')
        else:
            flash_form_errors(request, form)
    else:
        form = StudentGroupForm()
    
    return render(request, 'classroom_core/group_form.html', {
        'form': form,
        'title': 'Создать группу'
    })

@login_required
def group_edit(request, group_id):
    """Редактирование группы студентов"""
    group = get_object_or_404(StudentGroup, id=group_id)
    
    if not(request.user == group.created_by or request.user.profile.is_staff() or request.user.is_superuser):
        raise PermissionDenied
    
    if request.method == 'POST':
        form = StudentGroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, 'Группа успешно обновлена')
            return redirect('classroom_core:group_list')
        else:
            flash_form_errors(request, form)
    else:
        form = StudentGroupForm(instance=group)
    
    return render(request, 'classroom_core/group_form.html', {
        'form': form,
        'title': 'Редактировать группу',
        'group': group
    })

@login_required
def group_delete(request, group_id):
    """Удаление группы студентов"""
    group = get_object_or_404(StudentGroup, id=group_id)
    
    if not(request.user == group.created_by or request.user.profile.is_staff() or request.user.is_superuser):
        raise PermissionDenied
    
    if request.method == 'POST':
        group_name = group.name
        group.delete()
        messages.success(request, f'Группа "{group_name}" успешно удалена')
        return redirect('classroom_core:group_list')
    
    return render(request, 'classroom_core/group_confirm_delete.html', {
        'group': group
    })

@login_required
def group_detail(request, group_id):
    """Детали группы студентов"""
    group = get_object_or_404(StudentGroup, id=group_id)
    
    students = group.students.all()
    
                         
    paginator = Paginator(students, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'group': group,
        'page_obj': page_obj,
    }
    
    return render(request, 'classroom_core/group_detail.html', context)

@login_required
def group_add_students(request, group_id):
    """Добавление студентов в группу"""
    group = get_object_or_404(StudentGroup, id=group_id)

    if not(request.user == group.created_by or request.user.profile.is_staff() or request.user.is_superuser):
        raise PermissionDenied

    if request.method == 'POST':
        student_ids = request.POST.getlist('students')
        students_added = 0

        for student_id in student_ids:
            try:
                user = User.objects.get(id=student_id)
                if hasattr(user, 'profile') and user.profile.role == 'student':
                    user.profile.student_group = group
                    user.profile.save()
                    students_added += 1
            except User.DoesNotExist:
                continue

        messages.success(request, f'В группу добавлено {students_added} студентов')
        return redirect('classroom_core:group_detail', group_id=group.id)

                                                 
    available_students = User.objects.filter(
        profile__role='student',
        profile__student_group__isnull=True
    )

    context = {
        'group': group,
        'available_students': available_students,
    }

    return render(request, 'classroom_core/group_add_students.html', context)


                                                                               
                          
                                                                               

@login_required
def course_enrollment_request_create(request, course_id):
    """Создание заявки студента на запись на курс"""
    course = get_object_or_404(Course, id=course_id)

                                                       
    if course.students.filter(id=request.user.id).exists():
        messages.error(request, 'Вы уже записаны на этот курс')
        return redirect('classroom_core:course_detail', course_id=course.id)

                                 
    existing_request = CourseEnrollmentRequest.objects.filter(
        course=course,
        student=request.user
    ).first()

    if existing_request:
        messages.info(request, 'У вас уже есть заявка на этот курс')
        return redirect('classroom_core:course_enrollment_request_detail', request_id=existing_request.id)

    if request.method == 'POST':
        form = CourseEnrollmentRequestForm(request.POST)
        if form.is_valid():
            request_obj = form.save(commit=False)
            request_obj.course = course
            request_obj.student = request.user
            request_obj.save()

            messages.success(request, 'Заявка на запись на курс успешно создана')
            return redirect('classroom_core:course_detail', course_id=course.id)
        else:
            flash_form_errors(request, form)
    else:
        form = CourseEnrollmentRequestForm()

    return render(request, 'classroom_core/course_enrollment_request_form.html', {
        'form': form,
        'course': course,
    })

@login_required
def for_enrollment_course_list(request):
    """Список всех курсов с возможностью записи для студента"""
    courses = Course.objects.filter(status='active')
    print(courses)
    user_requests = CourseEnrollmentRequest.objects.filter(
        student=request.user
    ).select_related('course')

    requests_dict = {}
    for req in user_requests:
        requests_dict[req.course_id] = req

    courses_with_status = []
    for course in courses:
        if course.students.filter(id=request.user.id).exists():
            continue
        
        enrollment_request = requests_dict.get(course.id)

        if enrollment_request is None:
            enrollment_status = 'not_enrolled'
            enrollment_request_id = None
        elif enrollment_request.status == 'pending':
            enrollment_status = 'pending'
            enrollment_request_id = enrollment_request.id
        elif enrollment_request.status == 'approved':
            enrollment_status = 'approved'
            enrollment_request_id = enrollment_request.id
        elif enrollment_request.status == 'rejected':
            enrollment_status = 'rejected'
            enrollment_request_id = enrollment_request.id
        else:
            enrollment_status = 'not_enrolled'
            enrollment_request_id = None
        
        courses_with_status.append({
            'course': course,
            'enrollment_status': enrollment_status,
            'enrollment_request_id': enrollment_request_id,
        })
    
    paginator = Paginator(courses_with_status, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    
    return render(request, 'classroom_core/for_enrollment_course_list.html', context)

@login_required
def course_enrollment_request_list(request, course_id):
    """Список заявок на запись на курс(для преподавателя)"""
    course = get_object_or_404(Course, id=course_id)

                   
    if not course.can_edit(request.user):
        raise PermissionDenied

    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all()
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())

    status = request.GET.get('status')
    requests = CourseEnrollmentRequest.objects.filter(course=course)

    if status:
        requests = requests.filter(status=status)

    requests = requests.order_by('-created_at')

    paginator = Paginator(requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'course': course,
        'page_obj': page_obj,
        'status': status,
        'is_teacher': is_teacher,
        'is_assistant': is_assistant,
        'is_admin': is_admin,
    }

    return render(request, 'classroom_core/course_enrollment_request_list.html', context)

@login_required
def course_enrollment_request_detail(request, request_id):
    """Детали заявки на запись на курс"""
    enrollment_request = get_object_or_404(CourseEnrollmentRequest, id=request_id)
    course = enrollment_request.course

                   
    if not enrollment_request.can_review(request.user) and enrollment_request.student != request.user:
        raise PermissionDenied

    is_student_user = enrollment_request.student == request.user
    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all() if hasattr(course, 'teaching_assistants') else False
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())

    context = {
        'enrollment_request': enrollment_request,
        'course': course,
        'is_student_user': is_student_user,
        'is_teacher': is_teacher,
        'is_assistant': is_assistant,
        'is_admin': is_admin,
    }

    return render(request, 'classroom_core/course_enrollment_request_detail.html', context)


@login_required
def course_enrollment_request_review(request, request_id):
    """Рассмотрение заявки на запись на курс(для преподавателя)"""
    enrollment_request = get_object_or_404(CourseEnrollmentRequest, id=request_id)
    course = enrollment_request.course

    if not enrollment_request.can_review(request.user):
        raise PermissionDenied

    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all() if hasattr(course, 'teaching_assistants') else False
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())

    if request.method == 'POST':
        form = CourseEnrollmentReviewForm(request.POST, instance=enrollment_request)
        if form.is_valid():
            enrollment_request = form.save(commit=False)
            enrollment_request.reviewed_by = request.user
            enrollment_request.reviewed_at = timezone.now()
            enrollment_request.save()

            if enrollment_request.status == 'approved':
                course.add_student(enrollment_request.student)
                messages.success(request, f'Заявка одобрена. Студент {enrollment_request.student.username} зачислен на курс')
            else:
                messages.info(request, f'Заявка отклонена. Статус: {enrollment_request.get_status_display()}')

            return redirect('classroom_core:course_enrollment_request_list', course_id=course.id)
        else:
            flash_form_errors(request, form)
    else:
        form = CourseEnrollmentReviewForm(instance=enrollment_request)

    return render(request, 'classroom_core/course_enrollment_request_review.html', {
        'form': form,
        'enrollment_request': enrollment_request,
        'course': course,
        'is_teacher': is_teacher,
        'is_assistant': is_assistant,
        'is_admin': is_admin,
    })


                                                                               
               
                                                                               

@login_required
def assignment_file_create(request, assignment_id):
    """Старая ссылка: единая форма отправки решения."""
    return redirect("classroom_core:assignment_submit", assignment_id=assignment_id)


@login_required
def assignment_file_list(request, assignment_id):
    """Список файлов задания(для преподавателя)"""
    assignment = get_object_or_404(Assignment, id=assignment_id)
    course = assignment.course

                   
    if not assignment.can_grade(request.user):
        raise PermissionDenied

    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all() if hasattr(course, 'teaching_assistants') else False
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())

    files = AssignmentFile.objects.filter(assignment=assignment).order_by('-uploaded_at')

    context = {
        'assignment': assignment,
        'course': course,
        'files': files,
        'is_teacher': is_teacher,
        'is_assistant': is_assistant,
        'is_admin': is_admin,
    }

    return render(request, 'classroom_core/assignment_file_list.html', context)


@login_required
def assignment_file_delete(request, file_id):
    """Удаление файла задания"""
    file_obj = get_object_or_404(AssignmentFile, id=file_id)
    assignment = file_obj.assignment
    course = assignment.course

                   
    if not file_obj.can_delete(request.user):
        raise PermissionDenied

    if request.method == 'POST':
        file_obj.delete()
        messages.success(request, 'Файл успешно удален')
        return redirect('classroom_core:assignment_detail', assignment_id=assignment.id)

    return render(request, 'classroom_core/assignment_file_confirm_delete.html', {
        'file': file_obj,
        'assignment': assignment,
        'course': course,
    })


                                                                               
                         
                                                                               

@login_required
def assignment_file_review_create(request, file_id):
    """Создание проверки файла задания"""
    file_obj = get_object_or_404(AssignmentFile, id=file_id)
    assignment = file_obj.assignment
    course = assignment.course

                   
    if not assignment.can_grade(request.user):
        raise PermissionDenied

    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all() if hasattr(course, 'teaching_assistants') else False
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())

                                                            
    existing_review = AssignmentFileReview.objects.filter(
        file=file_obj,
        reviewer=request.user
    ).first()

    if existing_review:
        messages.info(request, 'Вы уже проверяли этот файл')
        return redirect('classroom_core:assignment_file_review_edit', review_id=existing_review.id)

    if request.method == 'POST':
        form = AssignmentFileReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.file = file_obj
            review.reviewer = request.user
            review.save()

            messages.success(request, 'Проверка файла успешно создана')
            return redirect('classroom_core:assignment_file_list', assignment_id=assignment.id)
        else:
            flash_form_errors(request, form)
    else:
        form = AssignmentFileReviewForm()

    return render(request, 'classroom_core/assignment_file_review_form.html', {
        'form': form,
        'file': file_obj,
        'assignment': assignment,
        'course': course,
        'is_teacher': is_teacher,
        'is_assistant': is_assistant,
        'is_admin': is_admin,
    })


@login_required
def assignment_file_review_edit(request, review_id):
    """Редактирование проверки файла задания"""
    review = get_object_or_404(AssignmentFileReview, id=review_id)
    file_obj = review.file
    assignment = file_obj.assignment
    course = assignment.course

                   
    if not review.can_review(request.user):
        raise PermissionDenied

    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all() if hasattr(course, 'teaching_assistants') else False
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())

    if request.method == 'POST':
        form = AssignmentFileReviewForm(request.POST, instance=review)
        if form.is_valid():
            form.save()
            messages.success(request, 'Проверка файла успешно обновлена')
            return redirect('classroom_core:assignment_file_list', assignment_id=assignment.id)
        else:
            flash_form_errors(request, form)
    else:
        form = AssignmentFileReviewForm(instance=review)

    return render(request, 'classroom_core/assignment_file_review_form.html', {
        'form': form,
        'file': file_obj,
        'assignment': assignment,
        'course': course,
        'review': review,
        'is_teacher': is_teacher,
        'is_assistant': is_assistant,
        'is_admin': is_admin,
    })


@login_required
def assignment_file_review_detail(request, review_id):
    """Детали проверки файла задания"""
    review = get_object_or_404(AssignmentFileReview, id=review_id)
    file_obj = review.file
    assignment = file_obj.assignment
    course = assignment.course

                   
    if not assignment.can_grade(request.user) and file_obj.student != request.user:
        raise PermissionDenied

    is_student_user = file_obj.student == request.user
    is_teacher = course.instructor == request.user
    is_assistant = request.user in course.teaching_assistants.all() if hasattr(course, 'teaching_assistants') else False
    is_admin = request.user.is_superuser or(hasattr(request.user, 'profile') and request.user.profile.is_staff())

    context = {
        'review': review,
        'file': file_obj,
        'assignment': assignment,
        'course': course,
        'is_student_user': is_student_user,
        'is_teacher': is_teacher,
        'is_assistant': is_assistant,
        'is_admin': is_admin,
    }

    return render(request, 'classroom_core/assignment_file_review_detail.html', context)


@login_required
def course_gradebook(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if not course.can_access(request.user):
        raise PermissionDenied

    can_grade = course.can_edit(request.user)
    if not can_grade:
        raise PermissionDenied

    _ensure_year_schedule(course)
    group_by = request.GET.get("group_by", "none")
    students = list(course.get_all_enrolled_students())
    students = sorted(students, key=lambda s: (s.last_name or "", s.first_name or "", s.username))
    lessons_qs = course.lessons.all()
    if course.end_date:
        lessons_qs = lessons_qs.filter(lesson_date__lte=course.end_date)
    lessons = list(lessons_qs)
    assignments = list(
        course.assignments.filter(status="published", due_date__isnull=False).order_by("due_date", "id")
    )

    marks = LessonGrade.objects.filter(lesson__course=course).select_related("student", "lesson")
    marks_map = {(m.student_id, m.lesson_id): m for m in marks}

    assignments_by_date = {}
    for assignment in assignments:
        due_date = assignment.due_date.date()
        assignments_by_date.setdefault(due_date, []).append(assignment)

    journal_columns = []
    inserted_assignment_dates = set()
    for lesson in lessons:
        journal_columns.append({
            "key": f"lesson-{lesson.id}",
            "type": "lesson",
            "id": lesson.id,
            "date": lesson.lesson_date,
            "weekday": WEEKDAY_RU[lesson.lesson_date.weekday()],
            "number": lesson.lesson_number,
            "topic": lesson.topic,
        })
        day_assignments = assignments_by_date.get(lesson.lesson_date, [])
        if day_assignments and lesson.lesson_date not in inserted_assignment_dates:
            for assignment in day_assignments:
                journal_columns.append({
                    "key": f"assignment-{assignment.id}",
                    "type": "assignment",
                    "id": assignment.id,
                    "date": lesson.lesson_date,
                    "weekday": WEEKDAY_RU[lesson.lesson_date.weekday()],
                    "title": assignment.title,
                    "max_points": assignment.max_points,
                })
            inserted_assignment_dates.add(lesson.lesson_date)

    assignment_submissions = AssignmentSubmission.objects.filter(
        assignment__in=assignments,
        student__in=students
    ).select_related("assignment", "student")
    submissions_map = {(sub.student_id, sub.assignment_id): sub for sub in assignment_submissions}

    grade_matrix = []
    for student in students:
        student_group = getattr(getattr(student, "profile", None), "student_group", None)
        row = {"student": student, "group_name": student_group.name if student_group else "Без группы", "cells": []}
        assignment_scores = []
        nb_count = 0
        for column in journal_columns:
            if column["type"] == "lesson":
                record = marks_map.get((student.id, column["id"]))
                mark_value = record.mark if record else ""
                row["cells"].append({
                    "type": "lesson",
                    "lesson_id": column["id"],
                    "mark": mark_value,
                })
                if isinstance(mark_value, str) and mark_value.strip().lower() in {"нб", "nb"}:
                    nb_count += 1
            else:
                submission = submissions_map.get((student.id, column["id"]))
                row["cells"].append({
                    "type": "assignment",
                    "assignment_id": column["id"],
                    "score": submission.score if submission and submission.score is not None else "",
                    "feedback": submission.feedback if submission else "",
                    "status": submission.status if submission else "",
                    "max_points": column["max_points"],
                })
                if submission and submission.score is not None:
                    assignment_scores.append(submission.score)
        row["assignment_avg"] = round(sum(assignment_scores) / len(assignment_scores), 2) if assignment_scores else None
        row["nb_count"] = nb_count
        grade_matrix.append(row)

    if group_by == "group":
        grade_matrix.sort(key=lambda item: (item["group_name"], item["student"].username))

    return render(
        request,
        "classroom_core/course_gradebook.html",
        {
            "course": course,
            "journal_columns": journal_columns,
            "grade_matrix": grade_matrix,
            "can_grade": can_grade,
            "group_by": group_by,
        },
    )


@login_required
@require_http_methods(["POST"])
def course_gradebook_add_lesson(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if not course.can_edit(request.user):
        raise PermissionDenied

    lesson_date_raw = request.POST.get("lesson_date")
    topic = (request.POST.get("topic") or "").strip()
    try:
        lesson_date = datetime.strptime(lesson_date_raw, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        messages.error(request, "Неверный формат даты пары")
        return redirect("classroom_core:course_gradebook", course_id=course.id)

    max_number = (
        CourseLesson.objects.filter(course=course, lesson_date=lesson_date)
        .order_by("-lesson_number")
        .values_list("lesson_number", flat=True)
        .first()
        or 0
    )
    lesson = CourseLesson.objects.create(
        course=course,
        lesson_date=lesson_date,
        lesson_number=max_number + 1,
        topic=topic,
    )
    messages.success(request, "Пара добавлена в журнал")
    return redirect(f"{reverse('classroom_core:course_gradebook', kwargs={'course_id': course.id})}?scroll_to_lesson={lesson.id}")


@login_required
@require_http_methods(["POST"])
def course_gradebook_update_topic(request, course_id, lesson_id):
    course = get_object_or_404(Course, id=course_id)
    if not course.can_edit(request.user):
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    lesson = get_object_or_404(CourseLesson, id=lesson_id, course=course)
    lesson.topic = (request.POST.get("topic") or "").strip()
    lesson.save(update_fields=["topic"])
    return JsonResponse({"success": True})


@login_required
def course_gradebook_update(request, course_id):
    """Обновление оценок в таблице оценок (AJAX POST)"""
    course = get_object_or_404(Course, id=course_id)

    if not course.can_edit(request.user):
        return JsonResponse({'success': False, 'error': 'Нет прав для редактирования'}, status=403)

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        assignment_id = request.POST.get('assignment_id')
        column_id = request.POST.get('column_id')
        lesson_id = request.POST.get("lesson_id")
        
        if not student_id or student_id == 'undefined':
            return JsonResponse({'success': False, 'error': 'Неверный ID студента'}, status=400)
        if (
            (not assignment_id or assignment_id == 'undefined')
            and (not column_id or column_id == 'undefined')
            and (not lesson_id or lesson_id == 'undefined')
        ):
            return JsonResponse({'success': False, 'error': 'Неверный ID объекта журнала'}, status=400)
        
        try:
            student_id = int(student_id)
            assignment_id = int(assignment_id) if assignment_id and assignment_id != 'undefined' else None
            column_id = int(column_id) if column_id and column_id != 'undefined' else None
            lesson_id = int(lesson_id) if lesson_id and lesson_id != "undefined" else None
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'ID должны быть числами'}, status=400)

        score = request.POST.get('score')
        feedback = request.POST.get('feedback', '')
        status = request.POST.get('status')
        mark = request.POST.get("mark")

        try:
            student = User.objects.get(id=student_id)
            assignment = Assignment.objects.get(id=assignment_id) if assignment_id else None
            column = GradebookColumn.objects.get(id=column_id, course=course) if column_id else None
            lesson = CourseLesson.objects.get(id=lesson_id, course=course) if lesson_id else None
        except (User.DoesNotExist, Assignment.DoesNotExist, GradebookColumn.DoesNotExist, CourseLesson.DoesNotExist):
            return JsonResponse({'success': False, 'error': 'Студент или объект оценки не найдены'}, status=404)

        if assignment and assignment.course != course:
            return JsonResponse({'success': False, 'error': 'Задание не принадлежит этому курсу'}, status=400)

        if lesson:
            lesson_grade, _ = LessonGrade.objects.get_or_create(lesson=lesson, student=student)
            normalized = (mark if mark is not None else score or "").strip()
            if normalized:
                lowered = normalized.lower()
                if lowered in {"нб", "nb"}:
                    normalized = "нб"
                else:
                    try:
                        int(normalized)
                    except ValueError:
                        return JsonResponse({'success': False, 'error': 'Разрешены только число или "нб"'}, status=400)
            lesson_grade.mark = normalized
            lesson_grade.feedback = feedback
            lesson_grade.graded_by = request.user
            lesson_grade.graded_at = timezone.now()
            lesson_grade.save()
            return JsonResponse({'success': True, 'message': 'Оценка в журнале обновлена'})
        elif assignment:
            submission, created = AssignmentSubmission.objects.get_or_create(
                assignment=assignment,
                student=student
            )
            if not created and submission.graded_at and (timezone.now() - submission.graded_at).days >= 3:
                return JsonResponse({'success': False, 'error': 'Редактирование доступно только в течение 3 дней после оценки'}, status=400)
        else:
            submission, created = GradebookRecord.objects.get_or_create(
                column=column,
                student=student
            )
            if not created and submission.graded_at and (timezone.now() - submission.graded_at).days >= 3:
                return JsonResponse({'success': False, 'error': 'Редактирование доступно только в течение 3 дней после оценки'}, status=400)
        if score is not None and score != '':
            try:
                submission.score = int(score)
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Оценка должна быть числом'}, status=400)
                
        submission.feedback = feedback
        if status:
            submission.status = status
            if status == 'graded':
                submission.graded_by = request.user
                submission.graded_at = timezone.now()

        submission.save()

        return JsonResponse({
            'success': True,
            'message': 'Оценка успешно обновлена'
        })

    return JsonResponse({'success': False, 'error': 'Метод не разрешен'}, status=405)


@login_required
def course_gradebook_export(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if not course.can_edit(request.user):
        raise PermissionDenied

    assignments = list(course.assignments.filter(status='published').order_by('created_at'))
    students = list(course.get_all_enrolled_students())
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments"

    header = ["student_id", "username", "full_name"] + [f"{a.id}:{a.title}" for a in assignments]
    ws.append(header)
    for student in students:
        row = [student.id, student.username, student.get_full_name()]
        for assignment in assignments:
            submission = AssignmentSubmission.objects.filter(assignment=assignment, student=student).first()
            row.append(submission.score if submission and submission.score is not None else "")
        ws.append(row)

    attendance_sheet = wb.create_sheet("Attendance")
    attendance_header = ["student_id", "username", "full_name"]
    lessons = list(course.lessons.order_by("lesson_date", "lesson_number"))
    attendance_header.extend([f"lesson:{lesson.id}:{lesson.lesson_date}" for lesson in lessons])
    attendance_sheet.append(attendance_header)
    lesson_marks = LessonGrade.objects.filter(lesson__course=course)
    marks_map = {(mark.student_id, mark.lesson_id): mark.mark for mark in lesson_marks}
    for student in students:
        row = [student.id, student.username, student.get_full_name()]
        for lesson in lessons:
            row.append(marks_map.get((student.id, lesson.id), ""))
        attendance_sheet.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="gradebook_course_{course.id}.xlsx"'
    return response


@login_required
@require_http_methods(["POST"])
def course_gradebook_import(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if not course.can_edit(request.user):
        raise PermissionDenied

    upload = request.FILES.get("gradebook_file")
    if not upload:
        messages.error(request, "Файл не загружен")
        return redirect("classroom_core:course_gradebook", course_id=course.id)

    workbook = load_workbook(upload)
    sheet = workbook["Assignments"] if "Assignments" in workbook.sheetnames else workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assignment_map = {}
    for col_idx, header in enumerate(headers[3:], start=4):
        if not header or ":" not in str(header):
            continue
        assignment_id = int(str(header).split(":", 1)[0])
        assignment = Assignment.objects.filter(id=assignment_id, course=course).first()
        if assignment:
            assignment_map[col_idx] = assignment

    updated = 0
    for row in sheet.iter_rows(min_row=2):
        student_id = row[0].value
        if not student_id:
            continue
        student = User.objects.filter(id=int(student_id)).first()
        if not student:
            continue
        for col_idx, assignment in assignment_map.items():
            score = row[col_idx - 1].value
            if score in (None, ""):
                continue
            submission, _ = AssignmentSubmission.objects.get_or_create(assignment=assignment, student=student)
            if submission.graded_at and (timezone.now() - submission.graded_at).days >= 3:
                continue
            submission.score = int(score)
            submission.status = "graded"
            submission.graded_by = request.user
            submission.graded_at = timezone.now()
            submission.save()
            updated += 1

    attendance_updated = 0
    if "Attendance" in workbook.sheetnames:
        attendance_sheet = workbook["Attendance"]
        attendance_headers = [cell.value for cell in next(attendance_sheet.iter_rows(min_row=1, max_row=1))]
        lesson_map = {}
        for col_idx, header in enumerate(attendance_headers[3:], start=4):
            raw = str(header or "")
            if raw.startswith("lesson:"):
                parts = raw.split(":")
                if len(parts) >= 2 and parts[1].isdigit():
                    lesson = CourseLesson.objects.filter(id=int(parts[1]), course=course).first()
                    if lesson:
                        lesson_map[col_idx] = lesson
        for row in attendance_sheet.iter_rows(min_row=2):
            student_id = row[0].value
            if not student_id:
                continue
            student = User.objects.filter(id=int(student_id)).first()
            if not student:
                continue
            for col_idx, lesson in lesson_map.items():
                raw_mark = row[col_idx - 1].value
                if raw_mark in (None, ""):
                    continue
                lesson_grade, _ = LessonGrade.objects.get_or_create(lesson=lesson, student=student)
                lesson_grade.mark = str(raw_mark).strip()
                lesson_grade.graded_by = request.user
                lesson_grade.graded_at = timezone.now()
                lesson_grade.save()
                attendance_updated += 1

    messages.success(request, f"Импорт завершен. Оценки: {updated}, посещаемость: {attendance_updated}")
    return redirect("classroom_core:course_gradebook", course_id=course.id)


@login_required
@require_http_methods(["POST"])
def assignment_quiz_submit(request, assignment_id):
    assignment = get_object_or_404(Assignment, id=assignment_id, assignment_type="quiz")
    course = assignment.course
    if not course.can_access(request.user) or not course.students.filter(id=request.user.id).exists():
        raise PermissionDenied
    if not assignment.can_submit():
        messages.error(request, "Тест недоступен для отправки")
        return redirect("classroom_core:assignment_detail", assignment_id=assignment.id)

    questions = assignment.quiz_questions.prefetch_related("options").all()
    all_correct = True
    total_questions = questions.count()
    correct_answers = 0
    answers_payload = []
    for question in questions:
        selected = set(request.POST.getlist(f"question_{question.id}"))
        question_options = list(question.options.all())
        correct = {str(option.id) for option in question_options if option.is_correct}
        question_correct = True
        if assignment.quiz_mode == "single" and len(selected) != 1:
            question_correct = False
        if selected != correct:
            question_correct = False
        if question_correct:
            correct_answers += 1
        all_correct = all_correct and question_correct
        answers_payload.append({
            "question_id": question.id,
            "question_text": question.question_text,
            "selected_option_ids": list(selected),
            "correct_option_ids": list(correct),
            "is_correct": question_correct,
            "options": [
                {
                    "id": option.id,
                    "text": option.option_text,
                    "is_correct": option.is_correct,
                    "is_selected": str(option.id) in selected,
                }
                for option in question_options
            ],
        })

    AssignmentQuizAttempt.objects.create(
        assignment=assignment,
        student=request.user,
        total_questions=total_questions,
        correct_answers=correct_answers,
        answers_payload=answers_payload,
        is_correct=all_correct,
    )
    score = 0
    if total_questions > 0:
        score = round((correct_answers / total_questions) * assignment.max_points)
    submission, _ = AssignmentSubmission.objects.get_or_create(
        assignment=assignment,
        student=request.user,
    )
    submission.score = score
    submission.status = "graded"
    submission.feedback = f"Верных ответов: {correct_answers}/{total_questions}"
    submission.graded_at = timezone.now()
    submission.graded_by = None
    submission.save()

    messages.success(request, f"Тест проверен: {correct_answers}/{total_questions}, балл: {score}/{assignment.max_points}")
    return redirect("classroom_core:assignment_detail", assignment_id=assignment.id)


@login_required
@require_http_methods(["POST"])
def course_gradebook_column_create(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if not course.can_edit(request.user):
        raise PermissionDenied
    title = request.POST.get("title", "").strip()
    column_type = request.POST.get("column_type", "custom")
    max_points = int(request.POST.get("max_points") or 100)
    if not title:
        messages.error(request, "Название колонки обязательно")
        return redirect("classroom_core:course_gradebook", course_id=course.id)
    GradebookColumn.objects.create(
        course=course,
        title=title,
        column_type=column_type,
        max_points=max_points,
        order=course.gradebook_columns.count(),
    )
    messages.success(request, "Колонка журнала добавлена")
    return redirect("classroom_core:course_gradebook", course_id=course.id)


@login_required
def custom_admin_dashboard(request):
    if not _can_access_management(request.user):
        raise PermissionDenied
    context = {
        "users_count": User.objects.count(),
        "courses_count": Course.objects.count(),
        "assignments_count": Assignment.objects.count(),
        "files_count": File.objects.count(),
        "pending_enrollments": CourseEnrollmentRequest.objects.filter(status="pending").count(),
        "is_super_management": _is_super_management(request.user),
    }
    return render(request, "classroom_core/admin_dashboard.html", context)


def _can_access_management(user):
    if not hasattr(user, "profile"):
        return user.is_superuser
    return user.is_superuser or user.profile.is_admin() or user.profile.is_staff() or user.profile.is_teacher()


def _is_super_management(user):
    if not hasattr(user, "profile"):
        return user.is_superuser
    return user.is_superuser or user.profile.is_admin() or user.profile.is_staff()


@login_required
def custom_admin_courses(request):
    if not _can_access_management(request.user):
        raise PermissionDenied
    if _is_super_management(request.user):
        courses = Course.objects.select_related("instructor").all().order_by("-created_at")
    else:
        courses = Course.objects.select_related("instructor").filter(
            Q(instructor=request.user) | Q(teaching_assistants=request.user)
        ).distinct().order_by("-created_at")
    return render(request, "classroom_core/admin_courses.html", {"courses": courses, "is_super_management": _is_super_management(request.user)})


@login_required
def custom_admin_course_create(request):
    if not _can_access_management(request.user):
        raise PermissionDenied
    form = ManagementCourseForm(request.POST or None)
    if not _is_super_management(request.user):
        form.fields["instructor"].queryset = User.objects.filter(id=request.user.id)
        form.fields["instructor"].initial = request.user
    if request.method == "POST":
        if form.is_valid():
            course = form.save()
            _ensure_year_schedule(course)
            messages.success(request, "Курс создан")
            return redirect("classroom_core:custom_admin_courses")
        flash_form_errors(request, form)
    return render(request, "classroom_core/admin_course_form.html", {"form": form, "title": "Создать курс"})


@login_required
def custom_admin_course_edit(request, course_id):
    if not _can_access_management(request.user):
        raise PermissionDenied
    course = get_object_or_404(Course, id=course_id)
    if not (_is_super_management(request.user) or course.instructor_id == request.user.id or course.teaching_assistants.filter(id=request.user.id).exists()):
        raise PermissionDenied
    form = ManagementCourseForm(request.POST or None, instance=course)
    if not _is_super_management(request.user):
        form.fields["instructor"].queryset = User.objects.filter(id=request.user.id)
    if request.method == "POST":
        if form.is_valid():
            updated_course = form.save()
            _ensure_year_schedule(updated_course)
            messages.success(request, "Курс обновлен")
            return redirect("classroom_core:custom_admin_courses")
        flash_form_errors(request, form)
    return render(request, "classroom_core/admin_course_form.html", {"form": form, "title": "Редактировать курс"})


@login_required
@require_http_methods(["POST"])
def custom_admin_course_delete(request, course_id):
    if not _can_access_management(request.user):
        raise PermissionDenied
    course = get_object_or_404(Course, id=course_id)
    if not (_is_super_management(request.user) or course.instructor_id == request.user.id):
        raise PermissionDenied
    course.delete()
    messages.success(request, "Курс удален")
    return redirect("classroom_core:custom_admin_courses")


@login_required
def custom_admin_assignments(request):
    if not _can_access_management(request.user):
        raise PermissionDenied
    assignments = Assignment.objects.select_related("course").all().order_by("-created_at")
    if not _is_super_management(request.user):
        assignments = assignments.filter(
            Q(course__instructor=request.user) | Q(course__teaching_assistants=request.user)
        ).distinct()
    return render(request, "classroom_core/admin_assignments.html", {"assignments": assignments, "is_super_management": _is_super_management(request.user)})


@login_required
def custom_admin_assignment_create(request):
    if not _can_access_management(request.user):
        raise PermissionDenied
    form = ManagementAssignmentForm(request.POST or None)
    if not _is_super_management(request.user):
        form.fields["course"].queryset = Course.objects.filter(
            Q(instructor=request.user) | Q(teaching_assistants=request.user)
        ).distinct()
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Задание создано")
            return redirect("classroom_core:custom_admin_assignments")
        flash_form_errors(request, form)
    return render(request, "classroom_core/admin_assignment_form.html", {"form": form, "title": "Создать задание"})


@login_required
def custom_admin_assignment_edit(request, assignment_id):
    if not _can_access_management(request.user):
        raise PermissionDenied
    assignment = get_object_or_404(Assignment, id=assignment_id)
    if not (_is_super_management(request.user) or assignment.course.instructor_id == request.user.id or assignment.course.teaching_assistants.filter(id=request.user.id).exists()):
        raise PermissionDenied
    form = ManagementAssignmentForm(request.POST or None, instance=assignment)
    if not _is_super_management(request.user):
        form.fields["course"].queryset = Course.objects.filter(
            Q(instructor=request.user) | Q(teaching_assistants=request.user)
        ).distinct()
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Задание обновлено")
            return redirect("classroom_core:custom_admin_assignments")
        flash_form_errors(request, form)
    return render(request, "classroom_core/admin_assignment_form.html", {"form": form, "title": "Редактировать задание"})


@login_required
@require_http_methods(["POST"])
def custom_admin_assignment_delete(request, assignment_id):
    if not _can_access_management(request.user):
        raise PermissionDenied
    assignment = get_object_or_404(Assignment, id=assignment_id)
    if not (_is_super_management(request.user) or assignment.course.instructor_id == request.user.id):
        raise PermissionDenied
    assignment.delete()
    messages.success(request, "Задание удалено")
    return redirect("classroom_core:custom_admin_assignments")


@login_required
def custom_admin_students(request):
    if not _can_access_management(request.user):
        raise PermissionDenied
    users = User.objects.select_related("profile").filter(profile__role="student").order_by("username")
    return render(request, "classroom_core/admin_students.html", {"users": users, "is_super_management": _is_super_management(request.user)})


@login_required
def custom_admin_student_create(request):
    if not _is_super_management(request.user):
        raise PermissionDenied
    form = ManagementUserForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Пользователь создан")
            return redirect("classroom_core:custom_admin_students")
        flash_form_errors(request, form)
    return render(request, "classroom_core/admin_student_form.html", {"form": form, "title": "Создать пользователя"})


@login_required
def custom_admin_student_edit(request, user_id):
    if not _is_super_management(request.user):
        raise PermissionDenied
    user_obj = get_object_or_404(User, id=user_id)
    form = ManagementUserForm(request.POST or None, instance=user_obj)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Пользователь обновлен")
            return redirect("classroom_core:custom_admin_students")
        flash_form_errors(request, form)
    return render(request, "classroom_core/admin_student_form.html", {"form": form, "title": "Редактировать пользователя"})


@login_required
@require_http_methods(["POST"])
def custom_admin_student_delete(request, user_id):
    if not _is_super_management(request.user):
        raise PermissionDenied
    user_obj = get_object_or_404(User, id=user_id)
    if user_obj.id == request.user.id:
        messages.error(request, "Нельзя удалить текущего пользователя")
        return redirect("classroom_core:custom_admin_students")
    user_obj.delete()
    messages.success(request, "Пользователь удален")
    return redirect("classroom_core:custom_admin_students")